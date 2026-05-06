import os
import re
import glob
import io
import json
import pandas as pd
from collections import defaultdict
from datetime import date, datetime


# ── Persistent config (circle heads + extra CC recipients) ──────────────
CONFIG_PATH = os.path.join("data", "reporting_config.json")

def _read_config():
    try:
        if os.path.exists(CONFIG_PATH):
            with open(CONFIG_PATH, encoding="utf-8") as f:
                return json.load(f)
    except Exception as e:
        print(f"[Config] Read error: {e}")
    return {}

def _write_config(cfg):
    os.makedirs("data", exist_ok=True)
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)

def get_active_circle_heads():
    """Returns circle heads from config file; falls back to hardcoded CIRCLE_HEADS."""
    cfg = _read_config()
    if cfg.get("circle_heads"):
        return {ch["circle"]: ch for ch in cfg["circle_heads"]}
    return CIRCLE_HEADS.copy()

def get_extra_recipients():
    """Returns persistent extra CC email list from config file."""
    return _read_config().get("extra_recipients", [])

def get_management_recipients():
    """Returns management team recipients from config.
    Falls back to empty list — configure via /REPORTING-CONFIG/MANAGEMENT."""
    return _read_config().get("management_recipients", [])

def get_test_mode():
    """Returns (is_test_mode, test_emails_list).
    When test_mode is True, ALL sends go to test_emails only."""
    cfg = _read_config()
    is_test = cfg.get("test_mode", True)
    # Prefer test_emails list; fall back to single test_email
    emails = cfg.get("test_emails")
    if not emails:
        single = cfg.get("test_email", "pranjalg.work@gmail.com")
        emails = [single] if single else ["pranjalg.work@gmail.com"]
    return is_test, emails

from services.email_service import send_email
from database import SessionLocal
from models import SiteMonitoring

LOGO_CID   = "logo@shaurrya"
LOGO_BYTES = None

def _load_logo():
    logo_path = os.path.normpath(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "logo.png")
    )
    if os.path.exists(logo_path):
        try:
            from PIL import Image
            img = Image.open(logo_path).convert("RGBA")
            w = 160
            img = img.resize((w, int(img.height * w / img.width)), Image.LANCZOS)
            buf = io.BytesIO()
            img.save(buf, "PNG", optimize=True, compress_level=9)
            print(f"[Logo] Loaded and compressed — {len(buf.getvalue())//1024} KB")
            return buf.getvalue()
        except Exception as e:
            print(f"[Logo] Error: {e}")
    return None

LOGO_BYTES = _load_logo()


def _send(recipients, subject, html, excel_bytes=None, excel_filename=None):
    imgs        = [(LOGO_CID, LOGO_BYTES, "png")] if LOGO_BYTES else None
    attachments = [(excel_filename, excel_bytes)] if excel_bytes and excel_filename else None
    send_email(recipients, subject, html, inline_images=imgs, attachments=attachments)


# =====================================================
# DAILY UPLOAD FOLDER — files uploaded via the portal
# =====================================================

DAILY_DIR = "data/daily"
os.makedirs(DAILY_DIR, exist_ok=True)

DAILY_FILES = {
    "employee":     os.path.join(DAILY_DIR, "employee.xlsx"),
    "attendance":   os.path.join(DAILY_DIR, "attendance.xlsx"),
    "distance":     os.path.join(DAILY_DIR, "distance.xlsx"),
    "forms":        os.path.join(DAILY_DIR, "forms.xlsx"),
    "managers":     os.path.join(DAILY_DIR, "managers.xlsx"),
    "forms_filled": os.path.join(DAILY_DIR, "forms_filled.xlsx"),
    "alarm":        os.path.join(DAILY_DIR, "alarm.csv"),
    "active_sites": os.path.join(DAILY_DIR, "active_sites.xlsx"),
}

# Legacy fallback paths (used only if daily files not uploaded yet)
LEGACY_FILES = {
    "distance":   "data/Distance Report -1st feb 25 to 30 Nov'25.xlsx",
    "employee":   "data/EMPLOYEE details'26.xlsx",
    "attendance": "data/Report-1773314624370.xlsx",
}

# =====================================================
# TEST RECIPIENTS — used when sending test emails
# Only these three addresses receive emails in test mode.
# =====================================================

TEST_RECIPIENTS = [
    "pranjalg.work@gmail.com",
]

# =====================================================
# CIRCLE HEAD CONFIGURATION
# Phone numbers are used to identify circle heads
# dynamically from whatever employee file is uploaded —
# no usernames are hard-coded here.
# Update only when circle heads change.
# =====================================================

CIRCLE_HEADS = {
    "Delhi":        {"head": "Saurabh Gupta",      "phone": "9990009220", "email": "saurabhgupta@shaurryatele.com",  "full_name": "Delhi",
                     "aliases": ["Delhi NCR", "DELHI NCR", "Gurgaon", "Gurugram", "Noida", "Faridabad", "Ghaziabad", "Delhi Ncr"]},
    "GJ":           {"head": "Rajnish Nimbark",     "phone": "7226080870", "email": "rajnish@shaurryatele.com",      "full_name": "Gujarat",
                     "aliases": ["Gujrat", "GJ", "Guj"]},
    "KA":           {"head": "Satish Megaraj",      "phone": "9538886655", "email": "satish.megaraj@shaurryatele.com","full_name": "Karnataka",
                     "aliases": ["KA", "Bangalore", "Bengaluru"]},
    "Maharashtra":  {"head": "Dattatray Ranmalkar", "phone": "8888806810", "email": "dattatray@shaurryatele.com",    "full_name": "Maharashtra",
                     "aliases": ["MH", "Maha"]},
    "Mumbai":       {"head": "Sunil Bhagwat",       "phone": "8108779091", "email": "sunilbhagwat@shaurryatele.com", "full_name": "Mumbai",
                     "aliases": ["MUMBAI", "Mum", "Bombay"]},
    "UPE":          {"head": "Deepanshu Pandey",    "phone": "9140864299", "email": "deepanshupandey@shaurryatele.com","full_name": "Uttar Pradesh East",
                     "aliases": ["UP East", "UPE", "U.P. East", "UP-East", "Uttar Pradesh East"]},
    "UPW":          {"head": "Rajesh Shukla",       "phone": "8826162006", "email": "rajesh.shukla@shaurryatele.com", "full_name": "Uttar Pradesh West",
                     "aliases": ["UP West", "UPW", "U.P. West", "UP-West", "Uttar Pradesh West"]},
    "WB & Kolkata": {"head": "Abhiman Ganguly",     "phone": "9903451369", "email": "abhiman.ganguly@shaurryatele.com","full_name": "West Bengal",
                     "aliases": ["WB", "Kolkata", "West Bengal", "WB & Kolkata"]},
    "MP & CG":      {"head": "Piyush Khobragade",   "phone": "9773459073", "email": "piyush.khobragade@shaurryatele.com","full_name": "Madhya Pradesh",
                     "aliases": ["MP", "CG", "MP & CG", "Madhya Pradesh", "Chhattisgarh", "MP and CG"]},
}


# =====================================================
# EMAIL STYLE — shared across all reports
# =====================================================

EMAIL_HEAD = """
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<style>
  body,table,td,p,a,span,div{-webkit-text-size-adjust:100%;-ms-text-size-adjust:100%;margin:0;padding:0}
  table,td{mso-table-lspace:0;mso-table-rspace:0;border-collapse:collapse}
  img{border:0;height:auto;line-height:100%;outline:none;text-decoration:none;-ms-interpolation-mode:bicubic;display:block}
  body{background:#E8E8E8;font-family:'Helvetica Neue',Helvetica,Arial,sans-serif}
  @media only screen and (max-width:620px){
    .email-container{width:100%!important;border-radius:0!important}
    .section-pad{padding:20px 16px!important}
    .logo-pad{padding:14px 16px!important}
    .hero-pad{padding:20px 16px 18px!important}
  }
</style>"""


# =====================================================
# HELPERS
# =====================================================

def _extract_alarm_names(raw_alarms):
    """Extract deduplicated alarm type names from multi-line Related Alarms text.
    e.g. '1. BTLV\\n   2026-04...\\n\\n2. L LVD Cut\\n...' → 'BTLV, L LVD Cut'
    """
    if not raw_alarms or not raw_alarms.strip():
        return ""
    names, seen = [], set()
    for line in raw_alarms.split("\n"):
        m = re.match(r"^\d+\.\s+(.+)", line.strip())
        if m:
            name = m.group(1).strip()
            if name and name not in seen:
                seen.add(name)
                names.append(name)
    return ", ".join(names) if names else ""


def get_latest_alarm_file():
    """Picks the most recently modified Alarm_Report CSV or SODN_Report xlsx in data/."""
    files = (
        glob.glob("data/Alarm_Report_*.csv") +
        glob.glob("data/SODN_Report_*.xlsx") +
        glob.glob("data/SODN_Report_*.csv")
    )
    if not files:
        print("[Report] No Alarm_Report/SODN_Report file found in data/")
        return None
    latest = max(files, key=os.path.getmtime)
    print(f"[Report] Using alarm/SODN file: {latest}")
    return latest


def attendance_badge(value):
    v = str(value).strip().lower()
    if v in ["present", "p", "yes"]:
        return f'<span class="badge badge-present">{value}</span>'
    if v in ["absent", "a", "no"]:
        return f'<span class="badge badge-absent">{value}</span>'
    return f'<span class="badge badge-na">{value}</span>'


def user_rows_html(users):
    rows = ""
    for u in users:
        dist = u['distance']
        try:
            dist_str = f"{float(dist):g}"   # 56.63 → "56.63", 136.0 → "136"
        except (ValueError, TypeError):
            dist_str = str(dist)
        rows += f"""
        <tr>
          <td><strong>{u['user']}</strong></td>
          <td>{attendance_badge(u['attendance'])}</td>
          <td>{dist_str} km</td>
          <td style="line-height:1.7">{u['form_names']}</td>
        </tr>"""
    return rows


def site_rows_html(sites):
    rows = ""
    for s in sites:
        rows += f"""
        <tr>
          <td>{s['site_id']}</td>
          <td>{s['site_name']}</td>
          <td><span class="badge badge-down">Down</span></td>
        </tr>"""
    return rows


def present_count(users):
    return sum(
        1 for u in users
        if str(u["attendance"]).strip().lower() in ["present", "p", "yes"]
    )


def absent_count(users):
    return sum(
        1 for u in users
        if str(u["attendance"]).strip().lower() in ["absent", "a", "no"]
    )


# =====================================================
# EMAIL BUILDING BLOCKS
# =====================================================

def _email_shell(title, subtitle, body_html, report_date=None):
    """Wraps content in the full branded email shell."""
    if LOGO_BYTES:
        logo_block = (
            f'<img src="cid:{LOGO_CID}" alt="Shaurrya Teleservices" '
            f'width="140" style="display:block;max-width:140px;height:auto;">'
        )
    else:
        logo_block = (
            '<span style="font-size:22px;font-weight:900;color:#CC0000;letter-spacing:1px;'
            'font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;text-transform:uppercase;">'
            'SHAURRYA</span>'
            '<span style="display:block;font-size:9px;font-weight:600;color:#999;'
            'letter-spacing:2px;text-transform:uppercase;margin-top:2px;">Teleservices Pvt. Ltd.</span>'
        )
    date_str = report_date if report_date else datetime.now().strftime("%d %B %Y")
    time_str = datetime.now().strftime("%I:%M %p")
    return f"""<!DOCTYPE html>
<html lang="en">
<head>{EMAIL_HEAD}</head>
<body style="margin:0;padding:0;background:#E8E8E8;font-family:'Helvetica Neue',Helvetica,Arial,sans-serif;">

<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#E8E8E8;">
<tr><td align="center" style="padding:30px 12px 24px;">

  <table class="email-container" width="600" cellpadding="0" cellspacing="0" border="0"
         style="max-width:600px;width:100%;background:#ffffff;border-radius:10px;
                overflow:hidden;box-shadow:0 2px 16px rgba(0,0,0,0.10);">

    <!-- ═══ LOGO BAR (white) ═══ -->
    <tr>
      <td class="logo-pad" style="background:#FFFFFF;padding:20px 28px;border-bottom:4px solid #CC0000;">
        <table width="100%" cellpadding="0" cellspacing="0" border="0">
          <tr>
            <td style="vertical-align:middle;">{logo_block}</td>
            <td align="right" style="vertical-align:middle;padding-left:12px;">
              <p style="margin:0;font-size:9.5px;font-weight:700;color:#AAAAAA;
                        text-transform:uppercase;letter-spacing:1.5px;line-height:1;">
                Daily Field Report
              </p>
              <p style="margin:5px 0 0;font-size:13px;font-weight:700;color:#111111;
                        letter-spacing:-0.2px;line-height:1;">
                {date_str}
              </p>
              <p style="margin:3px 0 0;font-size:11px;font-weight:400;color:#888888;line-height:1;">
                {time_str}
              </p>
            </td>
          </tr>
        </table>
      </td>
    </tr>

    <!-- ═══ HERO BAND (red) ═══ -->
    <tr>
      <td class="hero-pad"
          style="background:#CC0000;padding:28px 28px 26px;">
        <p style="margin:0;font-size:24px;font-weight:800;color:#FFFFFF;
                  letter-spacing:-0.6px;line-height:1.25;
                  font-family:'Helvetica Neue',Helvetica,Arial,sans-serif;">
          {title}
        </p>
        <p style="margin:9px 0 0;font-size:13px;font-weight:400;
                  color:rgba(255,255,255,0.72);line-height:1.5;letter-spacing:0.1px;">
          {subtitle}
        </p>
      </td>
    </tr>

    <!-- ═══ BODY ═══ -->
    {body_html}

    <!-- ═══ FOOTER ═══ -->
    <tr>
      <td style="background:#1A1A1A;padding:16px 28px;">
        <table width="100%" cellpadding="0" cellspacing="0" border="0">
          <tr>
            <td style="vertical-align:middle;">
              <p style="margin:0;font-size:11px;color:rgba(255,255,255,0.38);line-height:1.5;
                        font-family:'Helvetica Neue',Helvetica,Arial,sans-serif;">
                Auto-generated &nbsp;&bull;&nbsp;
                <span style="color:#CC0000;font-weight:600;">Shaurrya Field Reporting System</span>
                &nbsp;&bull;&nbsp; {date_str}
              </p>
            </td>
            <td align="right" style="vertical-align:middle;padding-left:12px;white-space:nowrap;">
              <p style="margin:0;font-size:10px;font-weight:600;color:rgba(255,255,255,0.22);
                        letter-spacing:0.8px;text-transform:uppercase;">Confidential</p>
            </td>
          </tr>
        </table>
      </td>
    </tr>

  </table>
</td></tr>
</table>
</body></html>"""


def _stat_cards(*cards):
    """
    Renders stat cards in a 2-column grid (pairs per row).
    This is structural — no CSS media queries needed, so it works on
    all email clients including Gmail mobile.
    4 cards → 2×2, 3 cards → 2 top + 1 full-width bottom, 2 cards → 1×2.
    """
    def _card(val, lbl, sep_right=False, sep_bottom=False, full_width=False):
        w = "100%" if full_width else "50%"
        br = "border-right:1px solid #ECECEC;" if sep_right else ""
        bb = "border-bottom:1px solid #ECECEC;" if sep_bottom else ""
        return (
            f'<td width="{w}" style="vertical-align:top;{br}{bb}">'
            f'<table width="100%" cellpadding="0" cellspacing="0" border="0">'
            f'<tr><td style="padding:26px 16px 24px;text-align:center;">'
            f'<div style="width:24px;height:3px;background:#CC0000;border-radius:2px;'
            f'margin:0 auto 12px;"></div>'
            f'<p style="margin:0;font-size:30px;font-weight:800;color:#111111;line-height:1;'
            f'letter-spacing:-1px;font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;">'
            f'{val}</p>'
            f'<p style="margin:8px 0 0;font-size:9px;font-weight:700;color:#CC0000;'
            f'text-transform:uppercase;letter-spacing:1.2px;'
            f'font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;">{lbl}</p>'
            f'</td></tr>'
            f'</table>'
            f'</td>'
        )

    n = len(cards)
    rows_html = ""
    for i in range(0, n, 2):
        last_row = (i + 2 >= n)
        if i + 1 < n:
            rows_html += (
                f'<tr>'
                f'{_card(cards[i][0], cards[i][1], sep_right=True, sep_bottom=not last_row)}'
                f'{_card(cards[i+1][0], cards[i+1][1], sep_right=False, sep_bottom=not last_row)}'
                f'</tr>'
            )
        else:
            rows_html += (
                f'<tr>'
                f'{_card(cards[i][0], cards[i][1], full_width=True)}'
                f'</tr>'
            )

    return (
        f'<tr><td style="background:#FFFFFF;border-top:1px solid #ECECEC;'
        f'border-bottom:1px solid #ECECEC;padding:0;">'
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0">'
        f'{rows_html}'
        f'</table>'
        f'</td></tr>'
    )


def _section_title(text):
    return (
        f'<table cellpadding="0" cellspacing="0" border="0" style="margin-bottom:14px;">'
        f'<tr>'
        f'<td style="width:4px;background:#CC0000;border-radius:2px;">&nbsp;</td>'
        f'<td style="padding-left:10px;font-size:14px;font-weight:700;color:#111111;'
        f'letter-spacing:-0.3px;font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;">'
        f'{text}</td>'
        f'</tr>'
        f'</table>'
    )


def _data_table(headers, rows_html):
    header_cells = "".join(
        f'<th style="background:#CC0000;color:#FFFFFF;padding:11px 16px;text-align:left;'
        f'font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.8px;'
        f'white-space:nowrap;font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;">{h}</th>'
        for h in headers
    )
    return (
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0"'
        f' style="border-collapse:collapse;overflow:hidden;border:1px solid #E8E8E8;'
        f'border-radius:8px;font-size:13px;font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;">'
        f'<thead><tr>{header_cells}</tr></thead>'
        f'<tbody>{rows_html}</tbody>'
        f'</table>'
    )


def _section_block(title_text, table_html, first=False):  # noqa: ARG001
    return (
        f'<tr><td style="padding:26px 28px 0;" class="section-pad">'
        f'{_section_title(title_text)}'
        f'{table_html}'
        f'</td></tr>'
        f'<tr><td style="padding:0 28px 4px;">'
        f'<div style="height:1px;background:#EEEEEE;margin-top:26px;"></div>'
        f'</td></tr>'
    )


def _divider_row():
    return '<tr><td style="padding:0 32px;"><div style="height:1px;background:#F5F0F0;"></div></td></tr>'


def _simple_notification_body(heading, details, note=None):
    """Minimal email body for Excel-only reports."""
    details_html = "".join(
        f'<tr><td style="padding:5px 0;font-size:13px;color:#555;{_F}">'
        f'<strong style="color:#111;">{k}:</strong>&nbsp; {v}</td></tr>'
        for k, v in details
    )
    note_html = ""
    if note:
        note_html = (
            f'<tr><td style="padding-top:18px;">'
            f'<div style="border-left:4px solid #CC0000;background:#FEF9F9;'
            f'padding:12px 16px;border-radius:4px;font-size:13px;color:#333;{_F}">'
            f'{note}</div></td></tr>'
        )
    return (
        f'<tr><td style="padding:28px 28px 26px;">'
        f'<p style="margin:0 0 18px;font-size:14px;font-weight:600;color:#111;{_F}">'
        f'{heading}</p>'
        f'<table cellpadding="0" cellspacing="0" border="0" width="100%">'
        f'{details_html}{note_html}'
        f'</table>'
        f'<p style="margin:22px 0 0;font-size:13px;color:#888;{_F}">'
        f'The full report is attached as an Excel file.</p>'
        f'</td></tr>'
    )


# =====================================================
# ROW BUILDERS
# =====================================================

_F = "font-family:'Helvetica Neue',Helvetica,Arial,sans-serif;"

def _user_rows(users):
    rows = ""
    for i, u in enumerate(users):
        bg = "#FFFFFF" if i % 2 == 0 else "#FAFAFA"
        dist = u['distance']
        try:
            dist_str = f"{float(dist):g} km"
        except (ValueError, TypeError):
            dist_str = str(dist)

        v = str(u['attendance']).strip().lower()
        if v in ["present", "p", "yes"]:
            badge = ('<span style="background:#EDF7EE;color:#2E7D32;padding:3px 11px;'
                     'border-radius:20px;font-size:11px;font-weight:700;letter-spacing:0.2px;">Present</span>')
        elif v in ["absent", "a", "no"]:
            badge = ('<span style="background:#FEF0F0;color:#CC0000;padding:3px 11px;'
                     'border-radius:20px;font-size:11px;font-weight:700;letter-spacing:0.2px;">Absent</span>')
        else:
            badge = (f'<span style="background:#F5F5F5;color:#777777;padding:3px 11px;'
                     f'border-radius:20px;font-size:11px;font-weight:700;">{u["attendance"]}</span>')

        rows += (
            f'<tr style="background:{bg};">'
            f'<td style="padding:12px 16px;border-bottom:1px solid #F0F0F0;color:#111111;'
            f'font-weight:600;{_F}">{u["user"]}</td>'
            f'<td style="padding:12px 16px;border-bottom:1px solid #F0F0F0;">{badge}</td>'
            f'<td style="padding:12px 16px;border-bottom:1px solid #F0F0F0;color:#444444;'
            f'white-space:nowrap;{_F}">{dist_str}</td>'
            f'<td style="padding:12px 16px;border-bottom:1px solid #F0F0F0;color:#555555;'
            f'line-height:1.7;{_F}">{u["form_names"]}</td>'
            f'</tr>'
        )
    return rows


def _site_rows(sites):
    rows = ""
    for i, s in enumerate(sites):
        bg = "#FFFFFF" if i % 2 == 0 else "#FAFAFA"
        rows += (
            f'<tr style="background:{bg};">'
            f'<td style="padding:12px 16px;border-bottom:1px solid #F0F0F0;color:#555555;'
            f'font-family:\'Courier New\',monospace;font-size:12px;">{s["site_id"]}</td>'
            f'<td style="padding:12px 16px;border-bottom:1px solid #F0F0F0;color:#111111;'
            f'font-weight:600;{_F}">{s["site_name"]}</td>'
            f'<td style="padding:12px 16px;border-bottom:1px solid #F0F0F0;"></td>'
            f'</tr>'
        )
    return rows


# keep old names as aliases so nothing else breaks
def attendance_badge(value):
    v = str(value).strip().lower()
    if v in ["present", "p", "yes"]:
        return f'<span style="background:#E8F5E9;color:#2E7D32;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:600;">{value}</span>'
    if v in ["absent", "a", "no"]:
        return f'<span style="background:#FFEBEE;color:#CC0000;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:600;">{value}</span>'
    return f'<span style="background:#F5F5F5;color:#757575;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:600;">{value}</span>'

def user_rows_html(users):  return _user_rows(users)
def site_rows_html(sites):  return _site_rows(sites)


# =====================================================
# DASHBOARD HELPERS
# =====================================================

def _parse_form_types(form_types_str):
    """Parse 'Form A × 3, Form B × 1' → {'Form A': 3, 'Form B': 1}"""
    result = {}
    if not form_types_str:
        return result
    for part in form_types_str.split(","):
        part = part.strip()
        if " × " in part:
            name, _, cnt_s = part.rpartition(" × ")
            try:
                result[name.strip()] = result.get(name.strip(), 0) + int(cnt_s.strip())
            except ValueError:
                pass
    return result


def _forms_by_circle(excel_rows):
    """Aggregate {circle: {form_name: count}} from excel_rows."""
    result = defaultdict(lambda: defaultdict(int))
    for row in excel_rows:
        circle = row.get("circle", "Other")
        for fname, cnt in _parse_form_types(row.get("form_types", "")).items():
            result[circle][fname] += cnt
    return {k: dict(v) for k, v in result.items()}


def _dashboard_body(total, present, absent, sites_down_count=None, forms_total=None, forms_by_circle=None, sites_health=None, absentees=None):
    """
    Dashboard body sections:
      1. Stats cards — always 2×2:
           management/circle: Total | Present | Absent | Sites Down
           manager:           Total | Present | Absent | Forms Filled
      2. Absentees list (when absentees provided)
      3. Forms Filled Today (by circle)
      4. Sites Health (circle: total vs down)
    """
    # Section 1: Stats — always 4 cards (2×2)
    if sites_down_count is not None:
        out = _stat_cards(
            (total,            "Total Employees"),
            (present,          "Present Today"),
            (absent,           "Absent Today"),
            (sites_down_count, "Sites Down Today"),
        )
    elif forms_total is not None:
        out = _stat_cards(
            (total,       "Total Employees"),
            (present,     "Present Today"),
            (absent,      "Absent Today"),
            (forms_total, "Forms Filled Today"),
        )
    else:
        out = _stat_cards(
            (total,   "Total Employees"),
            (present, "Present Today"),
            (absent,  "Absent Today"),
        )

    # Section 2: Absentees
    if absentees:
        show_circle = any(a.get("circle") for a in absentees)
        abs_rows = ""
        for i, a in enumerate(absentees):
            bg = "#FFFFFF" if i % 2 == 0 else "#FAFAFA"
            abs_rows += (
                f'<tr style="background:{bg};">'
                f'<td style="padding:10px 16px;border-bottom:1px solid #F5F5F5;color:#111;'
                f'font-weight:600;{_F}">{a["name"]}</td>'
            )
            if show_circle:
                abs_rows += (
                    f'<td style="padding:10px 16px;border-bottom:1px solid #F5F5F5;color:#555;{_F}">'
                    f'{a.get("circle","")}</td>'
                )
            abs_rows += '</tr>'
        hdrs = ["Full Name", "Circle"] if show_circle else ["Full Name"]
        out += _section_block("Absentees Today", _data_table(hdrs, abs_rows))

    # Section 3: Forms
    if forms_by_circle:
        form_rows_html = ""
        show_circle_hdrs = len(forms_by_circle) > 1
        for circle in sorted(forms_by_circle.keys()):
            forms = forms_by_circle[circle]
            if not forms:
                continue
            if show_circle_hdrs:
                form_rows_html += (
                    f'<tr style="background:#FDF2F2;">'
                    f'<td colspan="2" style="padding:7px 14px;font-size:10px;font-weight:700;'
                    f'color:#CC0000;text-transform:uppercase;letter-spacing:1px;{_F}">'
                    f'{circle}</td>'
                    f'</tr>'
                )
            for i, (fname, cnt) in enumerate(sorted(forms.items())):
                bg     = "#FFFFFF" if i % 2 == 0 else "#FAFAFA"
                indent = "padding-left:28px;" if show_circle_hdrs else "padding-left:16px;"
                form_rows_html += (
                    f'<tr style="background:{bg};">'
                    f'<td style="{indent}padding-top:8px;padding-bottom:8px;padding-right:14px;'
                    f'border-bottom:1px solid #F5F5F5;color:#333;font-size:13px;{_F}">{fname}</td>'
                    f'<td style="padding:8px 18px;border-bottom:1px solid #F5F5F5;text-align:center;'
                    f'font-weight:800;color:#CC0000;font-size:16px;width:60px;{_F}">{cnt}</td>'
                    f'</tr>'
                )
        if form_rows_html:
            out += _section_block("Forms Filled Today", _data_table(["Form Name", "Count"], form_rows_html))

    # Section 3: Sites Health
    if sites_health:
        sorted_circles = sorted(sites_health.keys())
        if len(sorted_circles) == 1:
            # Single-circle view — render as two mini number cards
            c       = sorted_circles[0]
            total_s = sites_health[c].get("total", 0)
            down_s  = sites_health[c].get("down", 0)
            num_color = "#CC0000" if down_s > 0 else "#2E7D32"
            lbl_color = "#CC0000" if down_s > 0 else "#2E7D32"
            mini_html = (
                f'<table width="100%" cellpadding="0" cellspacing="0" border="0">'
                f'<tr>'
                f'<td width="50%" style="text-align:center;padding:22px 16px;'
                f'border-right:1px solid #ECECEC;">'
                f'<div style="width:20px;height:3px;background:#888;border-radius:2px;'
                f'margin:0 auto 10px;"></div>'
                f'<p style="margin:0;font-size:28px;font-weight:800;color:#111111;'
                f'letter-spacing:-1px;{_F}">{total_s}</p>'
                f'<p style="margin:7px 0 0;font-size:9px;font-weight:700;color:#999;'
                f'text-transform:uppercase;letter-spacing:1.2px;{_F}">Total Sites</p>'
                f'</td>'
                f'<td width="50%" style="text-align:center;padding:22px 16px;">'
                f'<div style="width:20px;height:3px;background:{num_color};border-radius:2px;'
                f'margin:0 auto 10px;"></div>'
                f'<p style="margin:0;font-size:28px;font-weight:800;color:{num_color};'
                f'letter-spacing:-1px;{_F}">{down_s}</p>'
                f'<p style="margin:7px 0 0;font-size:9px;font-weight:700;color:{lbl_color};'
                f'text-transform:uppercase;letter-spacing:1.2px;{_F}">Down Today</p>'
                f'</td>'
                f'</tr>'
                f'</table>'
            )
            out += _section_block("Sites Health", mini_html)
        else:
            # Multi-circle view — full table
            site_rows_html = ""
            for circle in sorted_circles:
                total_s    = sites_health[circle].get("total", 0)
                down_s     = sites_health[circle].get("down", 0)
                durations  = [d for d in sites_health[circle].get("durations", []) if d]
                if down_s > 0:
                    status_html = (
                        f'<span style="display:inline-block;background:#FEF0F0;color:#CC0000;'
                        f'padding:3px 10px;border-radius:12px;font-weight:700;font-size:12px;{_F}">'
                        f'{down_s} down</span>'
                    )
                    if durations:
                        rows_inner = "".join(
                            f'<tr>'
                            f'<td style="padding:2px 8px 2px 0;font-size:11px;color:#555;white-space:nowrap;{_F}">{d["name"]}</td>'
                            f'<td style="padding:2px 0;font-size:11px;font-weight:700;color:#CC0000;white-space:nowrap;{_F}">{d["duration"]}</td>'
                            f'</tr>'
                            for d in durations
                        )
                        duration_html = (
                            f'<table cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">'
                            f'{rows_inner}</table>'
                        )
                    else:
                        duration_html = f'<span style="font-size:12px;color:#999;{_F}">—</span>'
                else:
                    status_html = (
                        f'<span style="display:inline-block;background:#F0FBF0;color:#2E7D32;'
                        f'padding:3px 10px;border-radius:12px;font-weight:700;font-size:12px;{_F}">'
                        f'All OK</span>'
                    )
                    duration_html = f'<span style="font-size:12px;color:#999;{_F}">—</span>'
                site_rows_html += (
                    f'<tr>'
                    f'<td style="padding:10px 16px;border-bottom:1px solid #F0F0F0;color:#111;'
                    f'font-weight:600;{_F}">{circle}</td>'
                    f'<td style="padding:10px 16px;border-bottom:1px solid #F0F0F0;text-align:center;'
                    f'color:#444;font-size:13px;{_F}">{total_s if total_s else "None"}</td>'
                    f'<td style="padding:10px 16px;border-bottom:1px solid #F0F0F0;text-align:center;">'
                    f'{status_html}</td>'
                    f'<td style="padding:10px 16px;border-bottom:1px solid #F0F0F0;text-align:center;">'
                    f'{duration_html}</td>'
                    f'</tr>'
                )
            if site_rows_html:
                _th = lambda txt, center=False: (
                    f'<th style="background:#CC0000;color:#FFFFFF;padding:11px 16px;'
                    f'text-align:{"center" if center else "left"};'
                    f'font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.8px;'
                    f'white-space:nowrap;font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;">{txt}</th>'
                )
                site_header = (
                    _th("Circle") + _th("Total Sites") + _th("Status", center=True)
                    + _th("Site Down Duration", center=True)
                )
                site_table = (
                    f'<table width="100%" cellpadding="0" cellspacing="0" border="0"'
                    f' style="border-collapse:collapse;overflow:hidden;border:1px solid #E8E8E8;'
                    f'border-radius:8px;font-size:13px;font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;">'
                    f'<thead><tr>{site_header}</tr></thead>'
                    f'<tbody>{site_rows_html}</tbody>'
                    f'</table>'
                )
                out += _section_block("Sites Health", site_table)

    out += (
        '<tr><td style="padding:18px 28px 28px;">'
        f'<p style="margin:0;font-size:12px;color:#999;{_F}">'
        'The full detailed report is attached as an Excel file.</p>'
        '</td></tr>'
    )
    return out


# =====================================================
# REPORT BUILDERS
# =====================================================

def build_manager_email(manager, users, report_date, excel_rows=None):
    # Use excel_rows as the authoritative source when available — it's the same data
    # that drives the Excel attachment, so stats and Excel will always agree.
    if excel_rows:
        _att = lambda r: str(r.get("attendance", "")).strip().lower()
        total   = len(excel_rows)
        present = sum(1 for r in excel_rows if _att(r) in ["present", "p", "yes"])
        absent  = sum(1 for r in excel_rows if _att(r) in ["absent", "a", "no"])
        absentees = [
            {"name": r.get("full_name") or r.get("user", "")}
            for r in excel_rows if _att(r) in ["absent", "a", "no"]
        ]
    else:
        total   = len(users)
        present = present_count(users)
        absent  = absent_count(users)
        absentees = [
            {"name": u["user"]}
            for u in users
            if str(u.get("attendance", "")).strip().lower() in ["absent", "a", "no"]
        ]

    forms_by_circle = None
    forms_total = 0
    if excel_rows:
        mgr_forms = defaultdict(int)
        for row in excel_rows:
            for fname, cnt in _parse_form_types(row.get("form_types", "")).items():
                mgr_forms[fname] += cnt
        if mgr_forms:
            forms_by_circle = {"_team": dict(mgr_forms)}
            forms_total = sum(mgr_forms.values())

    body = _dashboard_body(total, present, absent, forms_total=forms_total, forms_by_circle=forms_by_circle, absentees=absentees or None)
    return _email_shell(
        "Daily Field Activity Report",
        f"Manager: <strong style='color:#fff;'>{manager}</strong> &nbsp;&middot;&nbsp; {report_date}",
        body,
        report_date=report_date,
    )


def build_circle_email(circle, head_name, users, sites, report_date, excel_rows=None, site_total=0):
    total   = len(users)
    present = present_count(users)
    absent  = absent_count(users)

    absentees = [
        {"name": u["user"]}
        for u in users
        if str(u.get("attendance", "")).strip().lower() in ["absent", "a", "no"]
    ]

    forms_by_circle = None
    if excel_rows:
        fbc = _forms_by_circle(excel_rows)
        if fbc:
            forms_by_circle = fbc

    down_count   = len(sites)
    total_count  = max(site_total, down_count)
    durations    = [
        {"name": s.get("site_name") or s.get("site_id") or "Unknown", "duration": s["duration"]}
        for s in sites if s.get("duration")
    ]
    sites_health = {circle: {"total": total_count, "down": down_count, "durations": durations}}

    body = _dashboard_body(
        total, present, absent,
        sites_down_count=down_count,
        forms_by_circle=forms_by_circle,
        sites_health=sites_health,
        absentees=absentees or None,
    )
    return _email_shell(
        f"Circle Daily Report — {circle}",
        f"Circle Head: <strong style='color:#fff;'>{head_name}</strong> &nbsp;&middot;&nbsp; {report_date}",
        body,
        report_date=report_date,
    )


_NOISE_CIRCLES = {"unknown", "nan", "none", ""}


def _normalize_circle(raw):
    """
    Dynamically map any circle/state name to its canonical name.
    Source of truth is CIRCLE_HEADS (canonical key + full_name).
    Resolution order:
      1. Exact match (case-insensitive) against canonical key or full_name
      2. Canonical key or full_name contained in raw (handles 'Delhi NCR' → 'Delhi')
      3. Raw contained in canonical key or full_name
      4. difflib fuzzy match against all known names (handles typos like 'Gujrat')
    """
    import difflib
    if not raw:
        return raw
    s = str(raw).strip()
    if not s or s.lower() in ("nan", "none", "unknown", ""):
        return s

    heads = get_active_circle_heads()
    s_lower = s.lower()

    # Build vocab: lowercased name → canonical key (canonical + full_name + aliases)
    vocab = {}
    for canon, info in heads.items():
        vocab[canon.lower()] = canon
        fn = info.get("full_name", "")
        if fn:
            vocab[fn.lower()] = canon
        for alias in info.get("aliases", []):
            if alias:
                vocab[str(alias).strip().lower()] = canon

    # 1. Exact match
    if s_lower in vocab:
        return vocab[s_lower]

    # 2 & 3. Substring containment
    for known_lower, canon in vocab.items():
        if known_lower in s_lower or s_lower in known_lower:
            return canon

    # 4. Fuzzy — match against all known names, pick best above threshold
    matches = difflib.get_close_matches(s_lower, list(vocab.keys()), n=1, cutoff=0.72)
    if matches:
        return vocab[matches[0]]

    return s

def build_management_email(management_data, site_down_data, report_date, excel_rows=None, site_total_data=None):
    total_staff      = sum(len(v) for v in management_data.values())
    total_present    = sum(present_count(v) for v in management_data.values())
    total_absent     = sum(absent_count(v) for v in management_data.values())
    total_sites_down = sum(len(v) for v in site_down_data.values())

    absentees = [
        {"name": u["user"], "circle": circle}
        for circle, users_list in management_data.items()
        for u in users_list
        if str(u.get("attendance", "")).strip().lower() in ["absent", "a", "no"]
    ]

    forms_by_circle = None
    if excel_rows:
        fbc = _forms_by_circle(excel_rows)
        if fbc:
            forms_by_circle = fbc

    # Include ALL circles: employee hierarchy + alarm file + DB — filter noise
    all_site_circles = set(
        list(management_data.keys()) +
        list(site_down_data.keys()) +
        (list(site_total_data.keys()) if site_total_data else [])
    )
    all_site_circles = {c for c in all_site_circles if c.strip().lower() not in _NOISE_CIRCLES}

    sites_health = None
    has_totals = bool(site_total_data)
    if all_site_circles:
        sites_health = {}
        for c in all_site_circles:
            down_sites = site_down_data.get(c, [])
            down_c  = len(down_sites)
            total_c = site_total_data.get(c) if has_totals else None
            if total_c is not None and total_c < down_c:
                total_c = down_c
            durations = [
                {"name": s.get("site_name") or s.get("site_id") or "Unknown", "duration": s["duration"]}
                for s in down_sites if s.get("duration")
            ]
            sites_health[c] = {"total": total_c, "down": down_c, "durations": durations}

    body = _dashboard_body(
        total_staff, total_present, total_absent,
        sites_down_count=total_sites_down,
        forms_by_circle=forms_by_circle,
        sites_health=sites_health,
        absentees=absentees or None,
    )
    return _email_shell(
        "All-Circle Daily Field Activity Report",
        f"Management Summary &nbsp;&middot;&nbsp; {report_date}",
        body,
        report_date=report_date,
    )


# =====================================================
# EXCEL REPORT GENERATOR
# =====================================================

def _validated_remark(attendance, forms_count):
    att = str(attendance).strip().lower()
    if att in ["a", "absent"]:       return "On Leave"
    if att in ["wfh"]:               return "WFH"
    if att in ["wfo"]:               return "WFO"
    if att in ["p", "present", "yes"]:
        return "Work done, Verified" if forms_count > 0 else "No work done"
    return str(attendance).strip() or "N/A"


def build_excel_report(rows, report_date, title="Productivity Report", sites_down=None, wfh_wfo_map=None):
    """
    rows: list of dicts with keys:
      full_name, username, circle, business_domain, role,
      attendance, distance, forms_count, form_types
    sites_down: optional list of dicts with keys: site_id, site_name, (circle optional)
    wfh_wfo_map: optional dict of username → "WFH" | "WFO" from GPS analysis
    Returns bytes of a styled .xlsx file with two or three sheets:
      1. Summary  (pivot: circle → employee → status counts)
      2. Productivity Report (per-employee detail)
      3. Sites Down (if sites_down provided)
    """
    _gps = {str(k).strip().lower(): v for k, v in (wfh_wfo_map or {}).items()}
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter
    import io

    BLUE  = "1E40AF"
    WHITE = "FFFFFF"
    LIGHT = "EFF6FF"
    GREY  = "F9FAFB"
    DARK  = "111111"

    thin  = Side(style="thin",   color="E5E7EB")
    thick = Side(style="medium", color="1E40AF")

    def hdr_fill():  return PatternFill("solid", fgColor=BLUE)
    def alt_fill(i): return PatternFill("solid", fgColor=LIGHT if i % 2 == 0 else WHITE)
    def grp_fill():  return PatternFill("solid", fgColor="DBEAFE")

    def hdr_font():  return Font(bold=True, color=WHITE, name="Calibri", size=10)
    def ttl_font():  return Font(bold=True, color=DARK,  name="Calibri", size=11)
    def bod_font():  return Font(color=DARK, name="Calibri", size=10)
    def grp_font():  return Font(bold=True, color=BLUE,  name="Calibri", size=10)

    def all_border(): return Border(left=thin, right=thin, top=thin, bottom=thin)
    def top_border(): return Border(left=thin, right=thin, top=thick, bottom=thin)

    def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Productivity Report ──────────────────────────────────
    ws = wb.active
    ws.title = "Productivity Report"

    STATUS_COLS = ["Work done, Verified", "No work done", "On Leave", "WFH", "WFO"]
    ATT_COLS    = ["Work done, Verified", "No work done", "On Leave"]
    GPS_COLS    = ["WFH", "WFO"]
    HEADERS = ["Full Name", "Username", "Circle", "Business Domain", "Role",
               "Attendance", "Distance (km)", "No of Forms Filled",
               "Form Type", "Validated Remark"]

    col_widths = [24, 18, 16, 18, 12, 12, 14, 18, 30, 22]

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    tc = ws["A1"]
    tc.value     = f"{title} — {report_date}"
    tc.font      = Font(bold=True, color=BLUE, name="Calibri", size=13)
    tc.alignment = center()
    tc.fill      = PatternFill("solid", fgColor="EFF6FF")
    ws.row_dimensions[1].height = 22

    # Header row
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font(); c.fill = hdr_fill()
        c.alignment = center(); c.border = all_border()
    ws.row_dimensions[2].height = 22

    # Data rows
    for ri, row in enumerate(rows, 3):
        vals = [
            row.get("full_name", ""),
            row.get("username", ""),
            row.get("circle", ""),
            row.get("business_domain", ""),
            row.get("role", ""),
            row.get("attendance", ""),
            row.get("distance", 0),
            row.get("forms_count", 0),
            row.get("form_types", ""),
            _validated_remark(row.get("attendance", ""), row.get("forms_count", 0)),
        ]
        fill = alt_fill(ri)
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = bod_font(); c.fill = fill
            c.alignment = left(); c.border = all_border()
        ws.row_dimensions[ri].height = 18

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A3"

    # ── Sheet 2: Summary ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")

    # Pivot: circle → {status: count}
    # Rules:
    #   GPS=WFH/WFO + Work done Verified → mark BOTH (work column + GPS column)
    #   GPS=WFH/WFO + No work done       → mark GPS column ONLY (suppress No work done)
    #   GPS=WFH/WFO + On Leave           → mark On Leave ONLY (GPS ignored)
    #   No GPS                           → mark attendance column as usual
    from collections import defaultdict
    circle_att  = defaultdict(lambda: defaultdict(int))  # attendance-based counts
    circle_gps  = defaultdict(lambda: defaultdict(int))  # GPS-based counts
    circle_emp  = defaultdict(int)                       # employee count per circle
    emp_data    = defaultdict(list)

    for row in rows:
        circle     = row.get("circle", "Other")
        uname      = str(row.get("username", "")).strip().lower()
        att_status = _validated_remark(row.get("attendance", ""), row.get("forms_count", 0))
        gps        = _gps.get(uname)  # "WFH", "WFO", or None

        circle_emp[circle] += 1

        if gps in ("WFH", "WFO") and att_status != "On Leave":
            if att_status == "Work done, Verified":
                circle_att[circle]["Work done, Verified"] += 1
            # "No work done" is suppressed when GPS is available
            circle_gps[circle][gps] += 1
        else:
            circle_att[circle][att_status] += 1

        emp_data[circle].append({
            "name":       row.get("full_name", row.get("username", "")),
            "att_status": att_status,
            "gps":        gps,
        })

    S_COLS = STATUS_COLS + ["Grand Total"]
    S_HEADERS = ["Row Labels"] + S_COLS

    # Title
    ws2.merge_cells(f"A1:{get_column_letter(len(S_HEADERS))}1")
    tc2 = ws2["A1"]
    tc2.value     = f"Employee Productivity Analysis — {report_date}"
    tc2.font      = Font(bold=True, color=BLUE, name="Calibri", size=13)
    tc2.alignment = center()
    tc2.fill      = PatternFill("solid", fgColor="EFF6FF")
    ws2.row_dimensions[1].height = 22

    # Column label row
    ws2.merge_cells(f"B2:{get_column_letter(len(S_HEADERS))}2")
    cl = ws2["B2"]
    cl.value = "Column Labels"; cl.font = hdr_font(); cl.fill = hdr_fill()
    cl.alignment = center(); cl.border = all_border()
    ws2.row_dimensions[2].height = 18

    # Header
    for ci, h in enumerate(S_HEADERS, 1):
        c = ws2.cell(row=3, column=ci, value=h)
        c.font = hdr_font(); c.fill = hdr_fill()
        c.alignment = center(); c.border = all_border()
    ws2.row_dimensions[3].height = 22

    cur_row = 4
    grand_att = defaultdict(int)
    grand_gps = defaultdict(int)
    grand_emp = 0

    for circle in sorted(circle_emp.keys()):
        ca      = circle_att[circle]
        cg      = circle_gps[circle]
        g_total = circle_emp[circle]
        grp_vals = (
            [circle]
            + [ca.get(s, "") or "" for s in ATT_COLS]
            + [cg.get(s, "") or "" for s in GPS_COLS]
            + [g_total]
        )
        for ci, val in enumerate(grp_vals, 1):
            c = ws2.cell(row=cur_row, column=ci, value=val if val != "" else None)
            c.font = grp_font(); c.fill = grp_fill()
            c.alignment = center() if ci > 1 else left()
            c.border = all_border()
        ws2.row_dimensions[cur_row].height = 18
        cur_row += 1

        for ei, emp in enumerate(emp_data[circle]):
            att   = emp["att_status"]
            gps   = emp["gps"]
            marks = defaultdict(int)

            if gps in ("WFH", "WFO") and att != "On Leave":
                if att == "Work done, Verified":
                    marks["Work done, Verified"] = 1
                # No work done suppressed — GPS column carries it
                marks[gps] = 1
            else:
                marks[att] = 1

            emp_vals = [emp["name"]] + [marks.get(s, "") or "" for s in STATUS_COLS] + [1]
            fill = alt_fill(ei)
            for ci, val in enumerate(emp_vals, 1):
                c = ws2.cell(row=cur_row, column=ci, value=val if val != "" else None)
                c.font = bod_font(); c.fill = fill
                c.alignment = center() if ci > 1 else left()
                c.border = all_border()
            ws2.row_dimensions[cur_row].height = 16
            cur_row += 1

        for s in ATT_COLS:
            grand_att[s] += ca.get(s, 0)
        for s in GPS_COLS:
            grand_gps[s] += cg.get(s, 0)
        grand_emp += g_total

    g_row = (
        ["Grand Total"]
        + [grand_att.get(s, "") or "" for s in ATT_COLS]
        + [grand_gps.get(s, "") or "" for s in GPS_COLS]
        + [grand_emp]
    )
    for ci, val in enumerate(g_row, 1):
        c = ws2.cell(row=cur_row, column=ci, value=val if val != "" else None)
        c.font = Font(bold=True, color=WHITE, name="Calibri", size=10)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = center() if ci > 1 else left()
        c.border = all_border()
    ws2.row_dimensions[cur_row].height = 20

    col_widths2 = [26, 20, 16, 14, 12, 10, 14]
    for ci, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.freeze_panes = "A4"

    # ── Sheet 3: Sites Down (optional, blue/white theme) ─────────────
    if sites_down is not None:
        SD_BLUE  = "1E40AF"
        SD_LIGHT = "EFF6FF"
        SD_WHITE = "FFFFFF"

        ws3 = wb.create_sheet("Sites Down")
        has_circle_col    = any("circle" in s for s in (sites_down or []))
        has_alarms_col    = any(s.get("alarms") for s in (sites_down or []))
        has_duration_col  = any(s.get("duration") for s in (sites_down or []))
        SD_HEADERS = []
        if has_circle_col:
            SD_HEADERS.append("Circle")
        SD_HEADERS += ["Site ID", "Site Name"]
        if has_alarms_col:
            SD_HEADERS.append("Status")
        if has_duration_col:
            SD_HEADERS.append("Duration")
        SD_HEADERS += ["SODN Start", "SODN End"]
        n_cols = len(SD_HEADERS)

        ws3.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        t = ws3["A1"]
        t.value = f"Sites Down — {report_date}"
        t.font = Font(bold=True, color=SD_BLUE, name="Calibri", size=13)
        t.alignment = center()
        t.fill = PatternFill("solid", fgColor=SD_LIGHT)
        ws3.row_dimensions[1].height = 22

        sd_thin   = Side(style="thin", color="BFDBFE")
        sd_border = Border(left=sd_thin, right=sd_thin, top=sd_thin, bottom=sd_thin)

        for ci, h in enumerate(SD_HEADERS, 1):
            c = ws3.cell(row=2, column=ci, value=h)
            c.font = Font(bold=True, color=SD_WHITE, name="Calibri", size=10)
            c.fill = PatternFill("solid", fgColor=SD_BLUE)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = sd_border
        ws3.row_dimensions[2].height = 22

        if not sites_down:
            ws3.merge_cells(f"A3:{get_column_letter(n_cols)}3")
            nc = ws3.cell(row=3, column=1, value="All sites operational — no outages reported.")
            nc.font = Font(bold=True, color="2E7D32", name="Calibri", size=11)
            nc.fill = PatternFill("solid", fgColor="E8F5E9")
            nc.alignment = Alignment(horizontal="center", vertical="center")
            ws3.row_dimensions[3].height = 20
        else:
            for ri, site in enumerate(sites_down, 3):
                row_fill = PatternFill("solid", fgColor=SD_LIGHT if ri % 2 == 0 else SD_WHITE)
                vals = []
                if has_circle_col:
                    vals.append(site.get("circle", ""))
                vals += [site.get("site_id", ""), site.get("site_name", "")]
                if has_alarms_col:
                    vals.append(site.get("alarms", ""))
                if has_duration_col:
                    vals.append(site.get("duration", ""))
                vals.append(site.get("sodn_start", ""))
                vals.append(site.get("sodn_end", ""))
                for ci, val in enumerate(vals, 1):
                    c = ws3.cell(row=ri, column=ci, value=val)
                    c.font = Font(color="111111", name="Calibri", size=10)
                    c.fill = row_fill
                    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    c.border = sd_border
                ws3.row_dimensions[ri].height = 18

        # Build column widths dynamically matching SD_HEADERS order
        sd_widths = []
        if has_circle_col:   sd_widths.append(16)   # Circle
        sd_widths += [22, 32]                        # Site ID, Site Name
        if has_alarms_col:   sd_widths.append(42)   # Status
        if has_duration_col: sd_widths.append(18)   # Duration
        sd_widths += [20, 20]                        # SODN Start, SODN End
        for ci, w in enumerate(sd_widths, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =====================================================
# CORE REPORT RUNNER — accepts explicit file paths
# =====================================================

def _find_col(columns, candidates):
    """Case-insensitive partial match across a list of candidate names.
    Columns are stringified first so datetime/int headers don't crash."""
    str_columns = [str(c) for c in columns]
    for candidate in candidates:
        for i, col_str in enumerate(str_columns):
            if candidate.lower() in col_str.lower():
                return columns[i]  # return the original column object
    return None


_WFH_WFO_PATH = os.path.join("data", "wfh_wfo_results.json")

def _load_wfh_wfo_map():
    """Returns username (lowercase) → 'WFH'|'WFO' from the last GPS analysis, or {}."""
    try:
        if os.path.exists(_WFH_WFO_PATH):
            with open(_WFH_WFO_PATH, encoding="utf-8") as f:
                data = json.load(f)
            return {
                str(r["username"]).strip().lower(): r["status"]
                for r in data.get("results", [])
                if r.get("status") in ("WFH", "WFO") and r.get("username")
            }
    except Exception as e:
        print(f"[WFH/WFO] Could not load GPS results: {e}")
    return {}


def _run_report(attendance_file, distance_file, employee_file, alarm_file=None,
                test_mode=False, send_types=None, extra_recipients=None, report_date=None):
    # send_types: subset of {"management", "circles", "managers"} — defaults to all three
    if send_types is None:
        send_types = {"management", "circles", "managers"}
    # Config-level test mode overrides the caller's test_mode flag
    _cfg_test_mode, _test_emails = get_test_mode()
    if _cfg_test_mode:
        test_mode = True
    _TEST_ADDR = _test_emails if _test_emails else TEST_RECIPIENTS
    # extra_recipients: ad-hoc emails added for this send only (merged with persistent config)
    _extra = list(set((extra_recipients or []) + get_extra_recipients()))
    # load circle heads from config (or hardcoded fallback)
    _circle_heads = get_active_circle_heads()
    # load GPS-based WFH/WFO map (username → status)
    _wfh_wfo_map = _load_wfh_wfo_map()
    if _wfh_wfo_map:
        print(f"[WFH/WFO] GPS map loaded: {len(_wfh_wfo_map)} employees")
    """Core logic: read files, merge data, build and send all emails."""

    if report_date:
        try:
            report_date = datetime.strptime(report_date, "%Y-%m-%d").strftime("%d %B %Y")
        except ValueError:
            report_date = datetime.now().strftime("%d %B %Y")
    else:
        report_date = datetime.now().strftime("%d %B %Y")

    # ---- Validate required files ----
    for f in [distance_file, employee_file, attendance_file]:
        if not os.path.exists(f):
            print(f"[Report] Missing file: {f} — aborting.")
            return {"success": False, "error": f"Missing file: {f}"}

    print("[Report] Reading data files...")
    distance_df   = pd.read_excel(distance_file)
    employee_df   = pd.read_excel(employee_file)
    attendance_df = pd.read_excel(attendance_file)

    # Force all column names to plain strings (Excel can parse headers as datetime/int)
    distance_df.columns   = distance_df.columns.astype(str).str.strip()
    employee_df.columns   = employee_df.columns.astype(str).str.strip()
    attendance_df.columns = attendance_df.columns.astype(str).str.strip()

    # ---- Load active employee whitelist ----
    _active_emp_path = os.path.join(os.path.dirname(__file__), "..", "data", "active_employees.json")
    _active_norm_usernames: set = set()
    try:
        with open(_active_emp_path, "r", encoding="utf-8") as _aef:
            _ae_data = json.load(_aef)
        for _ae in _ae_data.get("employees", []):
            u = str(_ae.get("username", "")).strip().lower()
            if u:
                # store both raw and timestamp-stripped forms
                _active_norm_usernames.add(u)
                _active_norm_usernames.add(re.sub(r'_\d{6,}$', '', u))
        print(f"[Report] Active employee whitelist loaded: {len(_ae_data.get('employees', []))} employees")
    except Exception as _aee:
        print(f"[Report] active_employees.json load error: {_aee} — no whitelist filtering applied")

    emp_cols = list(employee_df.columns)

    # ---- Auto-detect employee file columns (flexible matching) ----
    emp_username_col = _find_col(emp_cols, [
        "Field Executive Username", "FE Username", "Username", "User Name",
        "User ID", "Emp ID", "Employee ID",
    ]) or emp_cols[0]

    manager_col = _find_col(emp_cols, [
        "Reporting Manager", "Manager Name", "Manager", "Mgr",
        "Team Lead", "Supervisor", "Reported To",
    ])

    full_name_col = _find_col(emp_cols, [
        "Full Name", "Employee Name", "Emp Name", "Name",
    ])

    circle_col = _find_col(emp_cols, [
        "Circle", "Circle Name", "Telecom Circle",
    ])

    email_col = _find_col(emp_cols, [
        "Email", "Email ID", "Mail", "Email Address",
    ])

    business_domain_col = _find_col(emp_cols, [
        "Business Domain", "Business Unit", "Domain", "Business", "Segment",
    ])

    role_col = _find_col(emp_cols, [
        "Mode of Job", "Role", "Designation", "Position", "Job Title", "Grade",
    ])

    print(f"[Report] Employee schema — username: '{emp_username_col}', "
          f"manager: '{manager_col}', circle: '{circle_col}', "
          f"name: '{full_name_col}', email: '{email_col}'")

    # ---- Normalize usernames ----
    distance_df["Username"] = (
        distance_df["Username"].astype(str).str.strip().str.lower()
    )
    employee_df[emp_username_col] = (
        employee_df[emp_username_col].astype(str).str.strip().str.lower()
    )
    attendance_df["Username"] = (
        attendance_df["Username"].astype(str).str.strip().str.lower()
    )

    # ---- Detect attendance column IN attendance_df BEFORE any merge ----
    PA_VALUES = {"p", "a", "present", "absent", "yes", "no"}

    def find_attendance_col(df):
        """Find column with P/A values. Name-check first, then value scan."""
        by_name = _find_col(list(df.columns), [
            "Attendance", "Attn", "Status", "Mark", "Present", "Absent",
        ])
        if by_name:
            print(f"[Report] Attendance col found by name: '{by_name}'")
            return by_name
        best_col, best_score = None, -1
        for col in df.columns:
            if str(col).lower() in ("username", "name", "date"):
                continue
            sample = df[col].dropna().astype(str).str.strip().str.lower()
            score = int(sample.isin(PA_VALUES).sum())
            if score > best_score:
                best_score, best_col = score, col
        result = best_col if (best_col and best_score > 0) else df.columns[-1]
        print(f"[Report] Attendance col found by value scan: '{result}' (score={best_score})")
        return result

    def find_distance_col(df):
        """Find column with numeric KM values. Name-check first, then numeric scan."""
        by_name = _find_col(list(df.columns), [
            "Distance", "KM", "Km", "Travel", "Dist", "Total KM",
        ])
        if by_name:
            print(f"[Report] Distance col found by name: '{by_name}'")
            return by_name
        best_col, best_ratio = df.columns[-1], 0.0
        for col in df.columns:
            if str(col).lower() in ("username", "name", "date"):
                continue
            sample = df[col].dropna().astype(str).str.strip()
            numeric_count = sample.apply(
                lambda x: x.replace(".", "", 1).lstrip("-").isdigit()
            ).sum()
            ratio = numeric_count / max(len(sample), 1)
            if ratio > best_ratio:
                best_ratio, best_col = ratio, col
        print(f"[Report] Distance col found by numeric scan: '{best_col}' (ratio={best_ratio:.2f})")
        return best_col

    attendance_column = find_attendance_col(attendance_df)
    distance_column   = find_distance_col(distance_df)

    # ---- Build per-username lookup dicts DIRECTLY from source files ----
    # This avoids all merge column-collision issues completely.
    attendance_lookup = {}
    for _, r in attendance_df.iterrows():
        uname = str(r.get("Username", "")).strip().lower()
        val   = r.get(attendance_column)
        val_s = str(val).strip() if val is not None and str(val).lower() not in ["nan", "none", ""] else "N/A"
        if uname:
            attendance_lookup[uname] = val_s

    distance_lookup = {}
    for _, r in distance_df.iterrows():
        uname = str(r.get("Username", "")).strip().lower()
        val   = r.get(distance_column, 0)
        if pd.isna(val) or str(val).strip() in ["--", "nan", ""]:
            val = 0.0
        else:
            try:
                val = float(val)
            except Exception:
                val = 0.0
        if uname:
            distance_lookup[uname] = val

    print(f"[Report] Attendance lookup: {len(attendance_lookup)} entries | sample: {dict(list(attendance_lookup.items())[:3])}")
    print(f"[Report] Distance lookup  : {len(distance_lookup)} entries | sample: {dict(list(distance_lookup.items())[:3])}")

    # ---- Merge only employee info (name, circle, manager) ----
    employee_df_clean = employee_df.copy()
    employee_df_clean["_uname"] = employee_df_clean[emp_username_col]

    # Apply active-employee whitelist: drop anyone not in the approved list
    if _active_norm_usernames:
        before = len(employee_df_clean)
        employee_df_clean = employee_df_clean[
            employee_df_clean["_uname"].apply(
                lambda u: str(u).strip().lower() in _active_norm_usernames
                          or re.sub(r'_\d{6,}$', '', str(u).strip().lower()) in _active_norm_usernames
            )
        ].reset_index(drop=True)
        print(f"[Report] Whitelist filter: {before} → {len(employee_df_clean)} employees")

    # ---- Build username → manager lookup from managers.xlsx ----
    # Normalise usernames by stripping Skedulomatic long-number suffixes so that
    # "abc_st_1765365615044" and "abc_st" resolve to the same person.
    def _norm_uname(u):
        return re.sub(r'_\d{6,}$', '', str(u).strip().lower())

    username_to_manager         = {}   # norm_employee_uname → manager raw value (username or name)
    username_to_manager_display = {}   # norm_employee_uname → manager display name (when available)

    managers_file = DAILY_FILES.get("managers")
    if managers_file and os.path.exists(managers_file):
        try:
            mgr_df = pd.read_excel(managers_file)
            mgr_df.columns = mgr_df.columns.astype(str).str.strip()
            mgr_cols = list(mgr_df.columns)

            mgr_user_col = _find_col(mgr_cols, [
                "Field Executive Username", "FE Username", "Username",
                "User Name", "User ID", "Emp ID", "Employee ID",
            ])
            # Manager identifier column (may be username OR display name)
            mgr_name_col = _find_col(mgr_cols, [
                "Reporting Manager Username", "RM Username",
                "Reporting Manager", "Manager", "Mgr",
                "Team Lead", "Supervisor", "Reported To",
            ])
            # Separate display-name column (e.g. Skedulomatic "Reporting Manager Name")
            mgr_display_col = _find_col(mgr_cols, [
                "Reporting Manager Name", "Manager Full Name", "Manager Display Name",
                "RM Name", "Team Lead Name", "Supervisor Name",
            ])

            print(f"[Report] managers.xlsx — user col: '{mgr_user_col}', "
                  f"manager col: '{mgr_name_col}', display col: '{mgr_display_col}'")

            if mgr_user_col and mgr_name_col:
                for _, mrow in mgr_df.iterrows():
                    uname  = _norm_uname(mrow.get(mgr_user_col, ""))
                    mval   = str(mrow.get(mgr_name_col, "")).strip()
                    mdisp  = str(mrow.get(mgr_display_col, "")).strip() if mgr_display_col else ""
                    if uname and mval and mval.lower() not in ["nan", "none", ""]:
                        username_to_manager[uname] = mval
                        if mdisp and mdisp.lower() not in ["nan", "none", ""]:
                            username_to_manager_display[uname] = mdisp
                print(f"[Report] Manager lookup built: {len(username_to_manager)} entries "
                      f"({len(username_to_manager_display)} with display names)")
            else:
                print("[Report] managers.xlsx: could not detect username or manager column")
        except Exception as e:
            print(f"[Report] managers.xlsx read error: {e}")
    else:
        print("[Report] managers.xlsx not found — using employee file for manager info")


    # ---- Build forms lookup from forms_filled.xlsx ----
    # Key: normalized employee full name → list of (form_name, count)
    forms_lookup = {}
    forms_filled_path = DAILY_FILES.get("forms_filled")
    if forms_filled_path and os.path.exists(forms_filled_path):
        try:
            ff_df = pd.read_excel(forms_filled_path, dtype=str).fillna("")
            ff_df.columns = ff_df.columns.str.strip()
            # Use the latest date in the file
            if "Action Date" in ff_df.columns:
                latest_ff_date = ff_df["Action Date"].replace("", pd.NA).dropna().max()
                ff_df = ff_df[ff_df["Action Date"] == latest_ff_date]
            for _, frow in ff_df.iterrows():
                emp_name  = str(frow.get("Employee Full Name", "")).strip()
                form_name = str(frow.get("Form Name", "")).strip().strip("\t")
                try:
                    cnt = int(str(frow.get("Records COUNT", "1")).strip())
                except (ValueError, TypeError):
                    cnt = 1
                if emp_name and form_name:
                    key = emp_name.lower()
                    forms_lookup.setdefault(key, {})
                    forms_lookup[key][form_name] = forms_lookup[key].get(form_name, 0) + cnt
            print(f"[Report] forms_filled lookup built: {len(forms_lookup)} employees")
        except Exception as e:
            print(f"[Report] forms_filled.xlsx read error: {e}")
    else:
        print("[Report] forms_filled.xlsx not found — forms column will be empty")

    manager_data    = defaultdict(list)
    circle_data     = defaultdict(list)
    management_data = defaultdict(list)
    excel_rows      = []

    # Build username → {name, email} for all employees (managers are also employees in the same file)
    username_to_info = {}
    for _, erow in employee_df.iterrows():
        uname = str(erow.get(emp_username_col, "")).strip().lower()
        fname = ""
        if full_name_col:
            n = erow.get(full_name_col)
            fname = str(n).strip() if n and str(n).lower() not in ["nan", "none", ""] else ""
        email = ""
        if email_col:
            e = erow.get(email_col)
            email = str(e).strip() if e and str(e).lower() not in ["nan", "none", ""] else ""
        if uname and uname not in ["nan", "none"]:
            username_to_info[uname] = {"name": fname or uname, "email": email}
    print(f"[Report] username_to_info built: {len(username_to_info)} entries")

    # Normalised lookup (strips Skedulomatic timestamp suffixes) for manager name resolution
    username_to_info_norm = {_norm_uname(k): v for k, v in username_to_info.items()}

    # Dynamically identify circle head usernames by matching phone numbers from
    # CIRCLE_HEADS against the uploaded employee file — no usernames hard-coded.
    def _norm_phone(p):
        return re.sub(r"\D", "", str(p))

    phone_col = _find_col(emp_cols, ["Phone", "Mobile", "Contact", "Phone Number"])
    phone_to_circle = {_norm_phone(v["phone"]): k for k, v in _circle_heads.items() if v.get("phone")}

    def _scan_for_circle_heads(df, uname_col, ph_col):
        result = {}
        for _, erow in df.iterrows():
            uname = str(erow.get(uname_col, "")).strip().lower()
            phone = _norm_phone(erow.get(ph_col, ""))
            if phone and phone in phone_to_circle:
                result[uname] = phone_to_circle[phone]
        return result

    circle_head_unames = {}  # username → circle name (built from uploaded file)
    if phone_col:
        circle_head_unames = _scan_for_circle_heads(employee_df, emp_username_col, phone_col)
        print(f"[Report] Matched {len(circle_head_unames)} circle heads from employee.xlsx phones")

    # Fallback: scan managers.xlsx if employee.xlsx had no phone column or no matches
    if not circle_head_unames:
        mgr_file = DAILY_FILES.get("managers")
        if mgr_file and os.path.exists(mgr_file):
            try:
                fb_df = pd.read_excel(mgr_file, dtype=str).fillna("")
                fb_df.columns = fb_df.columns.astype(str).str.strip()
                fb_uname_col = _find_col(list(fb_df.columns), ["Field Executive Username", "Username"])
                fb_phone_col = _find_col(list(fb_df.columns), ["Phone", "Mobile", "Contact", "Phone Number"])
                if fb_uname_col and fb_phone_col:
                    circle_head_unames = _scan_for_circle_heads(fb_df, fb_uname_col, fb_phone_col)
                    print(f"[Report] Matched {len(circle_head_unames)} circle heads from managers.xlsx phones (fallback)")
            except Exception as e:
                print(f"[Report] managers.xlsx fallback scan error: {e}")

    if not circle_head_unames:
        print("[Report] WARNING: No circle heads matched — circle reports will not be sent")

    # Build parent map: employee username → reporting manager username (for circle lookup)
    parent_map = {}
    if manager_col:
        for _, erow in employee_df.iterrows():
            uname = str(erow.get(emp_username_col, "")).strip().lower()
            mgr   = str(erow.get(manager_col, "")).strip().lower()
            if uname and mgr and mgr not in ["nan", "none", ""]:
                parent_map[uname] = mgr
    if not parent_map and username_to_manager:
        # employee.xlsx has no Reporting Manager col — use the managers.xlsx mapping
        for uname, mgr_val in username_to_manager.items():
            parent_map[uname.lower()] = str(mgr_val).strip().lower()

    def find_circle(username, max_depth=8):
        """Walk up the reporting chain to find this employee's circle."""
        current = username
        for _ in range(max_depth):
            if current in circle_head_unames:
                return circle_head_unames[current]
            parent = parent_map.get(current)
            if not parent or parent == current:
                break
            current = parent
        return "Other"

    manager_email_map = {}  # manager display name → manager's actual email
    # Build canonical manager name+email lookup from reporting_config managers array.
    # This ensures Skedulomatic usernames (e.g. sanjeevg_st_1765365615044) always
    # resolve to their proper display name regardless of which format appears in the files.
    _canonical_mgr_name  = {}  # norm_uname → display name
    _canonical_mgr_email = {}  # norm_uname → email
    _cfg_managers = _read_config().get("managers", [])
    for _cm in _cfg_managers:
        _cm_email = str(_cm.get("email", "")).strip()
        _cm_name  = str(_cm.get("name", "")).strip()
        if _cm_email and _cm_name:
            # derive username from email local-part (e.g. sanjeevg_st_1765365615044@skedulomatic.com)
            _cm_local = _cm_email.split("@")[0]
            _cm_norm  = _norm_uname(_cm_local)
            _canonical_mgr_name[_cm_norm]  = _cm_name
            _canonical_mgr_email[_cm_norm] = _cm_email
    # Also seed from employee file (for managers who are also FEs)
    for _k, _v in username_to_info_norm.items():
        if _v.get("name") and _k not in _canonical_mgr_name:
            _canonical_mgr_name[_k] = _v["name"]
        if _v.get("email") and _k not in _canonical_mgr_email:
            _canonical_mgr_email[_k] = _v["email"]

    # Iterate over employee file — one row per employee
    for _, row in employee_df_clean.iterrows():
        username = str(row.get("_uname", "")).strip().lower()

        # Full name
        raw_name  = row.get(full_name_col) if full_name_col else None
        user_name = str(raw_name).strip() if raw_name and str(raw_name).lower() not in ["nan", "none", ""] else username

        # Circle — hierarchy walk is authoritative; only fall back to direct column
        # when the hierarchy can't resolve (returns "Other")
        circle = find_circle(username)
        if circle == "Other" and circle_col:
            cv = row.get(circle_col)
            if cv and str(cv).strip().lower() not in ["nan", "none", ""]:
                circle = _normalize_circle(str(cv).strip())

        # Manager — look up by normalised username (handles _st_TIMESTAMP suffix mismatches)
        norm_uname = _norm_uname(username)
        # managers.xlsx is the sole source of truth for manager mapping
        manager_raw = (username_to_manager.get(norm_uname) or
                       username_to_manager.get(username))

        # Resolve manager display name — priority:
        #   1. Dedicated "Reporting Manager Name" column from managers.xlsx
        #   2. Canonical name from reporting_config.json managers array (handles Skedulomatic IDs)
        #   3. Employee file lookup
        #   4. Fallback: normalized raw username (strips timestamp suffix)
        norm_raw = _norm_uname(manager_raw or "")
        manager_display = (username_to_manager_display.get(norm_uname) or
                           username_to_manager_display.get(username, "") or
                           _canonical_mgr_name.get(norm_raw, "") or
                           (username_to_info.get((manager_raw or "").lower()) or {}).get("name", "") or
                           (username_to_info_norm.get(norm_raw) or {}).get("name", ""))

        # Always use a normalized key so all timestamp-suffix variants collapse into one group
        manager = manager_display or norm_raw or circle

        # Populate email map — canonical config email takes priority
        if manager and manager not in manager_email_map:
            mgr_email_val = (_canonical_mgr_email.get(norm_raw) or
                             (username_to_info.get((manager_raw or "").lower()) or {}).get("email") or
                             (username_to_info_norm.get(norm_raw) or {}).get("email", ""))
            if mgr_email_val:
                manager_email_map[manager] = mgr_email_val

        # Attendance — direct lookup from attendance file (no merge ambiguity)
        attendance = attendance_lookup.get(username, "N/A")

        # Distance — direct lookup from distance file
        distance = distance_lookup.get(username, 0.0)

        # Forms — from forms_filled.xlsx, matched by employee full name
        emp_forms = forms_lookup.get(user_name.lower(), {})
        if not emp_forms:
            form_display = '<span style="color:#9e9e9e;font-size:12px">No forms</span>'
        else:
            form_display = "<br>".join(
                f'<span class="badge badge-ok">{fname}</span> ×{cnt}'
                for fname, cnt in sorted(emp_forms.items())
            )

        forms_count = sum(emp_forms.values()) if emp_forms else 0
        form_types  = ", ".join(
            f"{fname} × {cnt}" for fname, cnt in sorted(emp_forms.items())
        ) if emp_forms else ""

        biz_domain = ""
        if business_domain_col:
            bd = row.get(business_domain_col)
            biz_domain = str(bd).strip() if bd and str(bd).lower() not in ["nan", "none", ""] else ""

        role_val = ""
        if role_col:
            rv = row.get(role_col)
            role_val = str(rv).strip() if rv and str(rv).lower() not in ["nan", "none", ""] else ""

        user_record = {
            "user":       user_name,
            "attendance": attendance,
            "distance":   distance,
            "form_names": form_display,
        }

        excel_rows.append({
            "full_name":       user_name,
            "username":        username,
            "circle":          circle,
            "business_domain": biz_domain,
            "role":            role_val,
            "attendance":      attendance,
            "distance":        distance,
            "forms_count":     forms_count,
            "form_types":      form_types,
            "manager":         manager,
        })

        # Only add to manager_data if this resolves to a known manager from reporting_config.
        # Employees with no managers.xlsx entry (norm_raw == "" or not in canonical list)
        # would otherwise create spurious "Manager: Mumbai" groups.
        if norm_raw and norm_raw in _canonical_mgr_name:
            manager_data[manager].append(user_record)
        circle_data[circle].append(user_record)
        management_data[circle].append(user_record)

    # =====================================================
    # ALARM / SITE DOWN PROCESSING
    # =====================================================

    site_down_data = defaultdict(list)

    # Resolve alarm file: use provided, or fallback to legacy glob
    resolved_alarm = alarm_file
    if not resolved_alarm or not os.path.exists(resolved_alarm):
        resolved_alarm = get_latest_alarm_file()

    if resolved_alarm:
        # File may be an Excel file uploaded with a .csv extension — detect by magic bytes
        with open(resolved_alarm, "rb") as _f:
            _magic = _f.read(4)

        def _read_alarm(path, is_excel):
            """Read alarm/SODN file, auto-detecting title rows."""
            _site_id_hints   = ["global id", "site id", "global_id", "site_id"]
            _site_name_hints = ["site name", "site_name", "tower", "location"]
            def _has_data_cols(df):
                cols_lower = [str(c).strip().lower() for c in df.columns]
                return any(h in cols_lower for h in _site_id_hints + _site_name_hints)
            for hdr in (0, 1, 2):
                try:
                    df = (pd.read_excel(path, header=hdr, dtype=str) if is_excel
                          else pd.read_csv(path, header=hdr, encoding="latin-1", dtype=str))
                    df.columns = df.columns.str.strip()
                    if _has_data_cols(df):
                        return df
                except Exception:
                    pass
            return (pd.read_excel(path, dtype=str) if is_excel
                    else pd.read_csv(path, encoding="latin-1", dtype=str))

        alarm_df = _read_alarm(resolved_alarm, is_excel=(_magic[:2] == b"PK"))
        alarm_df = alarm_df.fillna("").astype(str)
        alarm_df.columns = alarm_df.columns.str.strip()

        # Detect columns dynamically
        a_cols         = list(alarm_df.columns)
        a_site_id_col  = _find_col(a_cols, ["Global ID", "Site ID", "Global_ID"])
        a_site_nm_col  = _find_col(a_cols, ["Site Name", "Site_Name", "Tower", "Location"])
        a_circle_col   = _find_col(a_cols, ["State/Circle", "State", "Circle", "Region", "Zone"])
        a_alarms_col   = _find_col(a_cols, ["Related Alarms", "Related_Alarms", "Alarms", "Alarm", "Reason", "Fault"])
        a_duration_col = _find_col(a_cols, [
            "Duration", "Down Duration", "Outage Duration", "Alarm Duration",
            "Down Time", "Downtime", "Duration (HH:MM:SS)", "Time", "Aging",
            "Down Since", "Outage Time", "Elapsed", "Hours Down",
        ])
        a_sodn_start_col = _find_col(a_cols, [
            "SODN Start", "Sodn Start", "Start Time", "Alarm Start", "Outage Start",
            "Open Time", "Opened At", "Event Start", "Fault Start", "Start",
        ])
        a_sodn_end_col = _find_col(a_cols, [
            "SODN End", "Sodn End", "End Time", "Alarm End", "Outage End",
            "Close Time", "Closed At", "Event End", "Fault End", "End",
        ])
        a_time_col = None
        if not a_duration_col:
            a_time_col = _find_col(a_cols, [
                "Alarm Start", "Start Time", "Created Time", "Created At", "Alarm Date",
                "First Alarm", "Outage Start", "Date Time", "DateTime", "Timestamp",
                "Alarm Time", "Created", "Date", "Alarm Raised Time", "Raised Time",
                "Event Time", "Fault Time", "Open Time", "Opened At", "Opened",
            ])
        print(f"[Report] Alarm cols — site_id: '{a_site_id_col}', site_name: '{a_site_nm_col}', "
              f"circle: '{a_circle_col}', alarms: '{a_alarms_col}', duration: '{a_duration_col}', "
              f"sodn_start: '{a_sodn_start_col}', sodn_end: '{a_sodn_end_col}', time: '{a_time_col}'")

        if a_site_id_col:
            alarm_df = alarm_df.drop_duplicates(subset=[a_site_id_col])

        db = SessionLocal()
        for _, row in alarm_df.iterrows():
            circle    = _normalize_circle(str(row.get(a_circle_col, "Unknown")).strip() if a_circle_col else "Unknown")
            site_id   = str(row.get(a_site_id_col, "Unknown")).strip() if a_site_id_col else "Unknown"
            site_name = str(row.get(a_site_nm_col, "Unknown")).strip() if a_site_nm_col else "Unknown"
            if not site_id or site_id in ("Unknown", "nan", ""):
                continue

            existing = db.query(SiteMonitoring).filter(
                SiteMonitoring.global_id == site_id
            ).first()

            if not existing:
                record = SiteMonitoring(
                    global_id=site_id,
                    site_name=site_name,
                    circle=circle,
                    status="Inactive",
                    alarm="Site Down",
                )
                db.add(record)

            raw_alarms     = str(row.get(a_alarms_col, "")).strip() if a_alarms_col else ""
            alarms_summary = _extract_alarm_names(raw_alarms)
            duration_val = str(row.get(a_duration_col, "")).strip() if a_duration_col else ""
            if duration_val.lower() in ("nan", "none", ""):
                duration_val = ""
            if not duration_val and a_time_col:
                raw_time = str(row.get(a_time_col, "")).strip()
                if raw_time and raw_time.lower() not in ("nan", "none", ""):
                    try:
                        alarm_dt = pd.to_datetime(raw_time, errors="coerce")
                        if alarm_dt is not pd.NaT and not pd.isnull(alarm_dt):
                            delta = datetime.now() - alarm_dt.to_pydatetime().replace(tzinfo=None)
                            total_secs = int(delta.total_seconds())
                            if total_secs > 0:
                                hours = total_secs // 3600
                                mins  = (total_secs % 3600) // 60
                                duration_val = f"{hours}h {mins}m" if hours > 0 else f"{mins}m"
                    except Exception:
                        pass
            sodn_start = str(row.get(a_sodn_start_col, "")).strip() if a_sodn_start_col else ""
            sodn_end   = str(row.get(a_sodn_end_col,   "")).strip() if a_sodn_end_col   else ""
            if sodn_start.lower() in ("nan", "none"): sodn_start = ""
            if sodn_end.lower()   in ("nan", "none"): sodn_end   = ""
            site_down_data[circle].append({
                "site_id":    site_id,
                "site_name":  site_name,
                "alarms":     alarms_summary,
                "duration":   duration_val,
                "sodn_start": sodn_start,
                "sodn_end":   sodn_end,
            })

        db.commit()
        db.close()
    else:
        print("[Report] Skipping alarm data — no alarm file found.")

    # Build site totals per circle — uploaded active-sites file takes priority over DB
    site_total_data = defaultdict(int)
    _active_sites_path = DAILY_FILES.get("active_sites", "")
    if _active_sites_path and os.path.exists(_active_sites_path):
        try:
            with open(_active_sites_path, 'rb') as _asf:
                _as_magic = _asf.read(4)
            if _as_magic[:2] == b'PK':
                as_df = pd.read_excel(_active_sites_path, dtype=str).fillna("")
            else:
                as_df = pd.read_csv(_active_sites_path, dtype=str, encoding='latin-1').fillna("")
            as_df.columns = as_df.columns.astype(str).str.strip()
            as_cols = list(as_df.columns)
            as_circle_col = _find_col(as_cols, ["Circle", "State/Circle", "State", "Region", "Zone"])
            as_count_col  = _find_col(as_cols, ["Total Sites", "Site Count", "Count", "Total", "Sites", "No of Sites", "Number of Sites"])
            if as_circle_col and as_count_col:
                for _, r in as_df.iterrows():
                    c = _normalize_circle(str(r.get(as_circle_col, "")).strip())
                    if not c or c.lower() in _NOISE_CIRCLES:
                        continue
                    try:
                        site_total_data[c] += int(str(r.get(as_count_col, 0)).strip())
                    except (ValueError, TypeError):
                        pass
                print(f"[Report] Active sites loaded from file: {dict(site_total_data)}")
            elif as_circle_col:
                # No count column — count rows per circle (one row = one site)
                for _, r in as_df.iterrows():
                    c = _normalize_circle(str(r.get(as_circle_col, "")).strip())
                    if c and c.lower() not in _NOISE_CIRCLES:
                        site_total_data[c] += 1
                print(f"[Report] Active sites counted by row: {dict(site_total_data)}")
            else:
                print("[Report] active_sites file: could not detect circle column")
        except Exception as _ase:
            print(f"[Report] active_sites file read error: {_ase}")
    else:
        # Fallback: derive totals from SiteMonitoring DB
        try:
            _db2 = SessionLocal()
            _all_db_sites = _db2.query(SiteMonitoring.circle, SiteMonitoring.global_id).all()
            _db2.close()
            for _sc, _ in _all_db_sites:
                site_total_data[_normalize_circle(str(_sc).strip())] += 1
            print(f"[Report] Active sites loaded from DB: {dict(site_total_data)}")
        except Exception as _ste:
            print(f"[Report] Could not query site totals from DB: {_ste}")

    # =====================================================
    # SEND — MANAGER REPORTS
    # =====================================================

    if "managers" in send_types:
        for manager, users in manager_data.items():
            mgr_rows = [r for r in excel_rows if r["manager"] == manager]
            body = build_manager_email(manager, users, report_date, excel_rows=mgr_rows)
            try:
                xl_bytes  = build_excel_report(mgr_rows, report_date, f"Manager Report — {manager}", wfh_wfo_map=_wfh_wfo_map)
                xl_name   = f"Manager_Report_{manager.replace(' ', '_')}_{report_date.replace(' ', '_')}.xlsx"
            except Exception as xe:
                print(f"[Report] Excel build failed for manager {manager}: {xe}")
                xl_bytes, xl_name = None, None
            if test_mode:
                recipients = _TEST_ADDR
                subject    = f"[TEST] Manager Report — {manager} | {report_date}"
            else:
                mgr_email  = manager_email_map.get(manager)
                recipients = list(set(([mgr_email] if mgr_email else []) + _extra)) or _TEST_ADDR
                subject    = f"[Daily Report] Manager — {manager} | {report_date}"
            _send(recipients, subject, body, excel_bytes=xl_bytes, excel_filename=xl_name)
            print(f"[Report] Manager report -> {manager} -> {recipients}")
        print("[Report] All manager reports sent.")

    # =====================================================
    # SEND — CIRCLE HEAD REPORTS
    # =====================================================

    if "circles" in send_types:
        for circle, users in circle_data.items():
            sites     = site_down_data.get(circle, [])
            ch        = _circle_heads.get(circle, {})
            head_name = ch.get("head", circle)
            circ_rows = [r for r in excel_rows if r["circle"] == circle]
            body      = build_circle_email(circle, head_name, users, sites, report_date,
                                           excel_rows=circ_rows,
                                           site_total=site_total_data.get(circle, 0))
            circ_sites = [{"site_id": s["site_id"], "site_name": s["site_name"],
                           "alarms": s.get("alarms", ""), "duration": s.get("duration", ""),
                           "sodn_start": s.get("sodn_start", ""), "sodn_end": s.get("sodn_end", "")}
                          for s in site_down_data.get(circle, [])]
            try:
                xl_bytes = build_excel_report(circ_rows, report_date, f"Circle Report — {circle}",
                                              sites_down=circ_sites, wfh_wfo_map=_wfh_wfo_map)
                xl_name  = f"Circle_Report_{circle.replace(' ', '_').replace('&','and')}_{report_date.replace(' ', '_')}.xlsx"
            except Exception as xe:
                print(f"[Report] Excel build failed for circle {circle}: {xe}")
                xl_bytes, xl_name = None, None
            if test_mode:
                recipients = _TEST_ADDR
                subject    = f"[TEST] Circle {circle} — {head_name} | {report_date}"
            else:
                ch_email   = ch.get("email")
                recipients = list(set(([ch_email] if ch_email else []) + _extra)) or _TEST_ADDR
                subject    = f"[Daily Report] Circle {circle} — {head_name} | {report_date}"
            _send(recipients, subject, body, excel_bytes=xl_bytes, excel_filename=xl_name)
            print(f"[Report] Circle report -> {circle} ({head_name}) -> {recipients}")
        print("[Report] Circle reports sent.")

    # =====================================================
    # SEND — MANAGEMENT REPORT
    # =====================================================

    if "management" in send_types:
        body = build_management_email(management_data, site_down_data, report_date,
                                      excel_rows=excel_rows, site_total_data=site_total_data)
        all_sites = [{"circle": c, "site_id": s["site_id"], "site_name": s["site_name"],
                      "alarms": s.get("alarms", ""), "duration": s.get("duration", ""),
                      "sodn_start": s.get("sodn_start", ""), "sodn_end": s.get("sodn_end", "")}
                     for c, sites in site_down_data.items() for s in sites]
        try:
            xl_bytes = build_excel_report(excel_rows, report_date, "All Circles Productivity Report",
                                          sites_down=all_sites, wfh_wfo_map=_wfh_wfo_map)
            xl_name  = f"Productivity_Report_All_{report_date.replace(' ', '_')}.xlsx"
        except Exception as xe:
            print(f"[Report] Excel build failed for management report: {xe}")
            xl_bytes, xl_name = None, None
        _mgmt_emails = [r["email"] for r in get_management_recipients() if r.get("email")]
        if test_mode:
            mgmt_recipients = _TEST_ADDR
            mgmt_subject    = f"[TEST] All Circles — Management Summary | {report_date}"
        else:
            mgmt_recipients = list(set(_mgmt_emails + _extra)) or _TEST_ADDR
            mgmt_subject    = f"[Daily Report] All Circles — Management Summary | {report_date}"
        _send(mgmt_recipients, mgmt_subject, body, excel_bytes=xl_bytes, excel_filename=xl_name)
        print("[Report] Management report sent.")

    return {"success": True}


# =====================================================
# PUBLIC ENTRY POINTS
# =====================================================

def send_daily_report():
    """
    Called by the scheduler (6 PM daily).
    Uses files uploaded via the portal (data/daily/).
    Falls back to legacy hardcoded paths if daily files are not present.
    """
    cfg = _read_config()
    if not cfg.get("mail_enabled", True):
        print("[Mail] Scheduler skipped — mail sending disabled via portal.")
        return
    att  = DAILY_FILES["attendance"]
    dist = DAILY_FILES["distance"]
    emp  = DAILY_FILES["employee"]
    alrm = DAILY_FILES["alarm"]

    # Fall back to legacy paths if daily uploads not found
    if not os.path.exists(att):
        att = LEGACY_FILES["attendance"]
    if not os.path.exists(dist):
        dist = LEGACY_FILES["distance"]
    if not os.path.exists(emp):
        emp = LEGACY_FILES["employee"]

    _run_report(
        attendance_file=att,
        distance_file=dist,
        employee_file=emp,
        alarm_file=alrm if os.path.exists(alrm) else None,
    )


def send_report_now(test_mode=False, send_types=None, extra_recipients=None, report_date=None):
    """
    Called by the API. send_types controls which email types are sent.
    extra_recipients: ad-hoc emails added only for this send (list of strings).
    report_date: YYYY-MM-DD string; if None defaults to today.
    """
    cfg = _read_config()
    if not cfg.get("mail_enabled", True):
        print("[Mail] Sending is disabled via portal kill switch — aborting.")
        return {"success": False, "error": "Mail sending is currently disabled. Enable it from the portal first."}

    missing = []
    for key in ["employee", "attendance", "distance", "forms_filled"]:
        if not os.path.exists(DAILY_FILES[key]):
            missing.append(key)

    if missing:
        return {
            "success": False,
            "error": f"Required files not uploaded yet: {', '.join(missing)}"
        }

    return _run_report(
        attendance_file=DAILY_FILES["attendance"],
        distance_file=DAILY_FILES["distance"],
        employee_file=DAILY_FILES["employee"],
        alarm_file=DAILY_FILES["alarm"] if os.path.exists(DAILY_FILES["alarm"]) else None,
        test_mode=test_mode,
        send_types=send_types,
        extra_recipients=extra_recipients,
        report_date=report_date,
    )


def get_recipients_preview():
    """
    Reads the currently uploaded employee + managers files and returns
    who would receive each email type — without sending anything.
    Used by the frontend confirmation modal.
    """
    emp_path = DAILY_FILES["employee"]
    if not os.path.exists(emp_path):
        return {"success": False, "error": "Employee file not uploaded yet"}

    try:
        employee_df = pd.read_excel(emp_path)
        employee_df.columns = employee_df.columns.astype(str).str.strip()
        emp_cols = list(employee_df.columns)

        emp_username_col = _find_col(emp_cols, [
            "Field Executive Username", "FE Username", "Username",
            "User Name", "User ID", "Emp ID", "Employee ID",
        ]) or emp_cols[0]
        manager_col  = _find_col(emp_cols, ["Reporting Manager", "Manager Name", "Manager", "Mgr", "Team Lead", "Supervisor"])
        full_name_col = _find_col(emp_cols, ["Full Name", "Employee Name", "Emp Name", "Name"])
        email_col     = _find_col(emp_cols, ["Email", "Email ID", "Mail", "Email Address"])
        phone_col     = _find_col(emp_cols, ["Phone", "Mobile", "Contact", "Phone Number"])

        employee_df[emp_username_col] = employee_df[emp_username_col].astype(str).str.strip().str.lower()

        # username → {name, email}
        username_to_info = {}
        for _, row in employee_df.iterrows():
            uname = str(row.get(emp_username_col, "")).strip().lower()
            fname = str(row.get(full_name_col, "")).strip() if full_name_col else ""
            email = str(row.get(email_col, "")).strip() if email_col else ""
            if fname.lower() in ["nan", "none", ""]: fname = uname
            if email.lower() in ["nan", "none", ""]: email = ""
            if uname and uname not in ["nan", "none"]:
                username_to_info[uname] = {"name": fname, "email": email}

        # Load managers.xlsx override
        username_to_manager = {}
        mgr_file = DAILY_FILES.get("managers", "")
        if mgr_file and os.path.exists(mgr_file):
            mgr_df = pd.read_excel(mgr_file)
            mgr_df.columns = mgr_df.columns.astype(str).str.strip()
            mgr_cols = list(mgr_df.columns)
            mu = _find_col(mgr_cols, ["Field Executive Username", "FE Username", "Username", "User Name"])
            mm = _find_col(mgr_cols, ["Reporting Manager", "Manager Name", "Manager"])
            if mu and mm:
                for _, mrow in mgr_df.iterrows():
                    u = str(mrow.get(mu, "")).strip().lower()
                    m = str(mrow.get(mm, "")).strip()
                    if u and m and m.lower() not in ["nan", "none", ""]:
                        username_to_manager[u] = m

        # ── Circle heads (phone-match against employee file) ──
        def _norm_phone(p): return re.sub(r"\D", "", str(p))
        _ch = get_active_circle_heads()
        phone_to_circle = {_norm_phone(v["phone"]): k for k, v in _ch.items() if v.get("phone")}

        circles_result = {}
        if phone_col:
            for _, row in employee_df.iterrows():
                phone = _norm_phone(row.get(phone_col, ""))
                if phone and phone in phone_to_circle:
                    circle = phone_to_circle[phone]
                    uname  = str(row.get(emp_username_col, "")).strip().lower()
                    info   = username_to_info.get(uname, {})
                    circles_result[circle] = {
                        "circle": circle,
                        "head":   info.get("name") or _ch[circle]["head"],
                        "email":  _ch[circle]["email"],
                    }

        # Fallback: use active circle heads if no phone matches
        if not circles_result:
            for circle, ch in _ch.items():
                circles_result[circle] = {"circle": circle, "head": ch["head"], "email": ch["email"]}

        # ── Managers (unique managers from employee + managers files) ──
        managers_result = {}
        for _, row in employee_df.iterrows():
            uname = str(row.get(emp_username_col, "")).strip().lower()
            mgr_val = username_to_manager.get(uname)
            if not mgr_val and manager_col:
                raw = row.get(manager_col)
                if raw and str(raw).lower() not in ["nan", "none", ""]:
                    mgr_val = str(raw).strip().lower()
            if mgr_val:
                info = username_to_info.get(mgr_val.lower(), {})
                name  = info.get("name") or mgr_val
                email = info.get("email", "")
                if name and name.lower() not in ["nan", "none", ""] and name not in managers_result:
                    managers_result[name] = email

        mgmt_list = get_management_recipients()
        if not mgmt_list:
            mgmt_list = [{"name": "Management Team", "email": "(not configured)"}]
        return {
            "success": True,
            "management": mgmt_list,
            "circles":    sorted(circles_result.values(), key=lambda x: x["circle"]),
            "managers":   [{"name": n, "email": e} for n, e in sorted(managers_result.items())],
        }

    except Exception as e:
        return {"success": False, "error": str(e)}


if __name__ == "__main__":
    send_daily_report()
